from rest_framework import viewsets, permissions
from rest_framework.response import Response
from rest_framework.decorators import action
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse
from datetime import datetime, timedelta
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth import logout, get_user_model
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from io import BytesIO
from .models import WaterQualityData, CustomUser, Reading, Machine, FlowMeterConfig
from .serializers import WaterQualityDataSerializer, ReadingSerializer
from .forms import CustomUserCreationForm
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from minimalmodbus import Instrument, NoResponseError, InvalidResponseError
from django.utils.timezone import now
import json
from django.contrib.auth.views import LoginView
from django.urls import reverse
from django.db.models import Sum  # Added for aggregating flow values

logger = logging.getLogger(__name__)
User = get_user_model()

class CustomLoginView(LoginView):
    template_name = 'registration/login.html'
    redirect_authenticated_user = True

    def get_success_url(self):
        return reverse('dashboard')

def is_admin(user):
    return user.is_authenticated and user.is_admin

@login_required
def dashboard(request):
    if request.user.is_admin:
        total_users = CustomUser.objects.filter(role=CustomUser.USER).count()
        recent_data_count = (WaterQualityData.objects.filter(timestamp__gte=timezone.now() - timedelta(days=7)).count() +
                            Reading.objects.filter(recorded_at__gte=timezone.now() - timedelta(days=7)).count())
        active_machines = Machine.objects.count()
        recent_activities = []
        return render(request, 'monitoring_app/admin_dashboard.html', {
            'total_users': total_users,
            'recent_data_count': recent_data_count,
            'active_machines': active_machines,
            'recent_activities': recent_activities
        })
    return render(request, 'monitoring_app/user_dashboard.html')

@login_required
def all_users_dashboard(request):
    if request.user.is_admin:
        users = CustomUser.objects.filter(role=CustomUser.USER)
        return render(request, 'monitoring_app/dashboard.html', {'users': users})
    return redirect('dashboard')

@login_required
@user_passes_test(is_admin)
def user_management(request):
    users = CustomUser.objects.all()
    return render(request, 'monitoring_app/user_management.html', {'users': users})

@login_required
@user_passes_test(is_admin)
def add_user(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            try:
                user = form.save(commit=False)
                password = f"{user.username}@123"
                user.set_password(password)
                user.save()
                for i in range(1, 11):
                    if getattr(user, f'show_flow{i}'):
                        machine_name = f'machine_flow_{i}'
                        machine, _ = Machine.objects.get_or_create(name=machine_name)
                        FlowMeterConfig.objects.get_or_create(
                            machine=machine,
                            defaults={
                                'port': '/dev/ttyUSB0',
                                'slave_id': i,
                                'flow_register': 2,
                                'total_register': 3
                            }
                        )
                messages.success(request, f'User {user.username} created successfully!')
                return redirect('user_management')
            except Exception as e:
                logger.error(f"Error saving user: {e}")
                messages.error(request, f"Error creating user: {str(e)}")
        else:
            logger.error(f"Form errors: {form.errors}")
            messages.error(request, "Please correct the errors below.")
    else:
        form = CustomUserCreationForm()
    return render(request, 'monitoring_app/add_user.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def delete_user(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    if request.method == 'POST':
        if user != request.user:
            user.delete()
            messages.success(request, f'User {user.username} deleted successfully!')
        else:
            messages.error(request, "You cannot delete your own account!")
        return redirect('user_management')
    return render(request, 'monitoring_app/delete_user_confirm.html', {'user': user})

@login_required
@user_passes_test(is_admin)
def edit_user(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            for i in range(1, 11):
                if getattr(user, f'show_flow{i}'):
                    machine_name = f'machine_flow_{i}'
                    machine, _ = Machine.objects.get_or_create(name=machine_name)
                    FlowMeterConfig.objects.get_or_create(
                        machine=machine,
                        defaults={
                            'port': '/dev/ttyUSB0',
                            'slave_id': i,
                            'flow_register': 2,
                            'total_register': 3
                        }
                    )
            messages.success(request, f'User {user.username} updated successfully!')
            return redirect('user_management')
        else:
            logger.error(f"Edit user form errors: {form.errors}")
    else:
        form = CustomUserCreationForm(instance=user)
    return render(request, 'monitoring_app/add_user.html', {'form': form, 'edit_mode': True})

@login_required
@user_passes_test(is_admin)
def calibrate_flow(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    if request.method == 'POST':
        try:
            flow_number = int(request.POST.get('flow_number'))
            new_total = float(request.POST.get('new_total'))
            if new_total < 0:
                messages.error(request, "Total flow cannot be negative.")
                return render(request, 'monitoring_app/calibrate_flow.html', {'user': user})
            machine = Machine.objects.get(name=f'machine_flow_{flow_number}')
            latest_reading = Reading.objects.filter(
                user=user, 
                machine=machine, 
                parameter=f'flow{flow_number}_total'
            ).order_by('-recorded_at').first()
            if latest_reading:
                latest_reading.value = new_total
                latest_reading.save()
                logger.info(f"Calibrated flow{flow_number}_total to {new_total} for user_id {user_id}")
            else:
                Reading.objects.create(
                    user=user,
                    machine=machine,
                    parameter=f'flow{flow_number}_total',
                    value=new_total
                )
                logger.info(f"Created new flow{flow_number}_total reading with value {new_total} for user_id {user_id}")
            messages.success(request, f'Total Flow {flow_number} calibrated successfully!')
            return redirect('user_management')
        except Machine.DoesNotExist:
            messages.error(request, f"Machine for Flow {flow_number} not found.")
        except ValueError as e:
            messages.error(request, f"Invalid input: {str(e)}")
        except Exception as e:
            messages.error(request, f"Error calibrating flow: {str(e)}")
    return render(request, 'monitoring_app/calibrate_flow.html', {'user': user})

@login_required
@user_passes_test(is_admin)
def fetch_external_flow(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    try:
        for flow_number in range(1, 11):
            if getattr(user, f'show_flow{flow_number}'):
                machine = Machine.objects.get(name=f'machine_flow_{flow_number}')
                try:
                    config = FlowMeterConfig.objects.get(machine=machine)
                except FlowMeterConfig.DoesNotExist:
                    logger.error(f"No FlowMeterConfig for machine_flow_{flow_number}")
                    messages.error(request, f"No configuration for Flow {flow_number}")
                    continue
                try:
                    instrument = Instrument(config.port, config.slave_id)
                    instrument.serial.baudrate = 9600
                    instrument.serial.timeout = 1
                    flow_value = instrument.read_float(config.flow_register)
                    flow_total = instrument.read_float(config.total_register)
                    if flow_value < 0 or flow_total < 0:
                        logger.error(f"Negative flow values for flow{flow_number}: flow={flow_value}, total={flow_total}")
                        messages.error(request, f"Negative flow values for Flow {flow_number}")
                        continue
                    today = now().date()
                    month_start = today.replace(day=1)
                    daily_readings = Reading.objects.filter(
                        user=user,
                        machine=machine,
                        parameter=f'flow{flow_number}',
                        recorded_at__date=today
                    )
                    monthly_readings = Reading.objects.filter(
                        user=user,
                        machine=machine,
                        parameter=f'flow{flow_number}',
                        recorded_at__date__gte=month_start
                    )
                    daily_flow = sum(reading.value for reading in daily_readings) + flow_value if daily_readings else flow_value
                    monthly_flow = sum(reading.value for reading in monthly_readings) + flow_value if monthly_readings else flow_value
                    Reading.objects.create(
                        user=user,
                        machine=machine,
                        parameter=f'flow{flow_number}',
                        value=flow_value
                    )
                    Reading.objects.create(
                        user=user,
                        machine=machine,
                        parameter=f'flow{flow_number}_total',
                        value=flow_total
                    )
                    Reading.objects.create(
                        user=user,
                        machine=machine,
                        parameter=f'flow{flow_number}_daily',
                        value=daily_flow
                    )
                    Reading.objects.create(
                        user=user,
                        machine=machine,
                        parameter=f'flow{flow_number}_monthly',
                        value=monthly_flow
                    )
                except NoResponseError:
                    logger.error(f"No response from flow meter {flow_number}")
                    messages.error(request, f"No response from Flow {flow_number}")
                    continue
                except InvalidResponseError:
                    logger.error(f"Invalid response from flow meter {flow_number}")
                    messages.error(request, f"Invalid response from Flow {flow_number}")
                    continue
                except Exception as e:
                    logger.error(f"Error reading flow meter {flow_number}: {e}")
                    messages.error(request, f"Error reading Flow {flow_number}: {str(e)}")
                    continue
        messages.success(request, 'External flow data fetched successfully!')
        return redirect('user_management')
    except Exception as e:
        logger.error(f"Error fetching external flow: {e}")
        messages.error(request, f"Error fetching external flow: {str(e)}")
        return redirect('user_management')

@login_required
@user_passes_test(is_admin)
def download_user_credentials(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    return generate_user_pdf(user)

def generate_user_pdf(user):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph("New User Credentials", styles['Title']))
    elements.append(Spacer(1, 12))
    user_data = [
        ['Username', user.username],
        ['Password', f"{user.username}@123"],
        ['Email', user.email],
        ['Role', user.get_role_display()],
        ['First Name', user.first_name],
        ['Last Name', user.last_name],
        ['Contact Number', user.contact_number or '-'],
        ['Address', user.address or '-'],
        ['State', user.state or '-'],
        ['State Code', user.state_code or '-']
    ]
    table = Table(user_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="user_{user.username}_credentials.pdf"'
    response.write(pdf)
    return response

class WaterQualityDataViewSet(viewsets.ModelViewSet):
    queryset = WaterQualityData.objects.all()
    serializer_class = WaterQualityDataSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
    
    def get_queryset(self):
        queryset = WaterQualityData.objects.all()
        user_id = self.request.query_params.get('user_id')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        limit = self.request.query_params.get('limit')
        if limit and limit.isdigit():
            queryset = queryset.order_by('-timestamp')[:int(limit)]
        else:
            queryset = queryset.order_by('-timestamp')
        return queryset
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        user_id = self.request.query_params.get('user_id')
        if user_id:
            try:
                user = CustomUser.objects.get(id=user_id)
                context['show_ph'] = user.show_ph
                context['show_cod'] = user.show_cod
                context['show_bod'] = user.show_bod
                context['show_tss'] = user.show_tss
            except CustomUser.DoesNotExist:
                logger.error(f"User with ID {user_id} not found in get_serializer_context")
        return context
    
    @action(detail=False, methods=['get'])
    def latest(self, request):
        user_id = request.query_params.get('user_id')
        if not user_id:
            logger.error("Missing user_id parameter in water quality latest action")
            return Response({"error": "user_id parameter is required"}, status=400)
        try:
            latest_data = WaterQualityData.objects.filter(user_id=user_id).order_by('-timestamp').first()
            if latest_data:
                serializer = self.get_serializer(latest_data)
                return Response(serializer.data)
            else:
                logger.info(f"No water quality data found for user_id {user_id}")
                return Response({
                    'timestamp': None,
                    'ph': None,
                    'cod': None,
                    'bod': None,
                    'tss': None
                })
        except Exception as e:
            logger.error(f"Error in water quality latest action for user_id {user_id}: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)
    
    @action(detail=False, methods=['get'])
    def date_range(self, request):
        user_id = request.query_params.get('user_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        time_interval = request.query_params.get('time_interval', '0')
        
        if not all([user_id, start_date, end_date]):
            logger.error("Missing required parameters in water quality date_range action")
            return Response({"error": "user_id, start_date, and end_date parameters are required"}, status=400)
        
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            time_interval = int(time_interval)
        except ValueError:
            logger.error(f"Invalid date format or time_interval in water quality date_range: start_date={start_date}, end_date={end_date}, time_interval={time_interval}")
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)
        
        if (end_date - start_date).days > 31:
            logger.error(f"Date range exceeds 31 days in water quality date_range: start_date={start_date}, end_date={end_date}")
            return Response({"error": "Date range cannot exceed 31 days"}, status=400)
        
        queryset = WaterQualityData.objects.filter(
            user_id=user_id,
            timestamp__date__gte=start_date,
            timestamp__date__lte=end_date
        ).order_by('timestamp')
        
        if time_interval > 0:
            interval_seconds = time_interval * 60
            filtered_data = []
            last_timestamp = None
            data_list = list(queryset)
            for item in data_list:
                if last_timestamp is None:
                    filtered_data.append(item)
                    last_timestamp = item.timestamp
                else:
                    time_diff = (item.timestamp - last_timestamp).total_seconds()
                    if time_diff >= interval_seconds:
                        filtered_data.append(item)
                        last_timestamp = item.timestamp
            queryset = filtered_data
        
        serializer = self.get_serializer(queryset, many=True)
        logger.info(f"Returning {len(serializer.data)} water quality records for user_id {user_id}")
        return Response(serializer.data)

class ReadingViewSet(viewsets.ModelViewSet):
    queryset = Reading.objects.all()
    serializer_class = ReadingSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
    
    def get_queryset(self):
        queryset = Reading.objects.all()
        user_id = self.request.query_params.get('user_id')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        limit = self.request.query_params.get('limit')
        if limit and limit.isdigit():
            queryset = queryset.order_by('-recorded_at')[:int(limit)]
        else:
            queryset = queryset.order_by('-recorded_at')
        return queryset
    
    @action(detail=False, methods=['get'])
    def latest(self, request):
        user_id = request.query_params.get('user_id')
        if not user_id:
            logger.error("Missing user_id parameter in readings latest action")
            return Response({"error": "user_id parameter is required"}, status=400)
        latest_readings = {}
        try:
            for i in range(1, 11):
                for param in [f'flow{i}', f'flow{i}_total', f'flow{i}_daily', f'flow{i}_monthly']:
                    reading = Reading.objects.filter(user_id=user_id, parameter=param).order_by('-recorded_at').first()
                    if reading:
                        latest_readings[param] = {
                            'value': reading.value,
                            'recorded_at': reading.recorded_at.isoformat()
                        }
            logger.info(f"Returning latest readings for user_id {user_id}")
            return Response(latest_readings)
        except Exception as e:
            logger.error(f"Error in readings latest action for user_id {user_id}: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)
    
    @action(detail=False, methods=['get'])
    def date_range(self, request):
        user_id = request.query_params.get('user_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        time_interval = request.query_params.get('time_interval', '0')
        
        if not all([user_id, start_date, end_date]):
            logger.error("Missing required parameters in readings date_range action")
            return Response({"error": "user_id, start_date, and end_date parameters are required"}, status=400)
        
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            time_interval = int(time_interval)
        except ValueError:
            logger.error(f"Invalid date format or time_interval in readings date_range: start_date={start_date}, end_date={end_date}, time_interval={time_interval}")
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)
        
        if (end_date - start_date).days > 31:
            logger.error(f"Date range exceeds 31 days in readings date_range: start_date={start_date}, end_date={end_date}")
            return Response({"error": "Date range cannot exceed 31 days"}, status=400)
        
        queryset = Reading.objects.filter(
            user_id=user_id,
            recorded_at__date__gte=start_date,
            recorded_at__date__lte=end_date
        ).order_by('recorded_at')
        
        if time_interval > 0:
            interval_seconds = time_interval * 60
            filtered_data = []
            last_timestamp = None
            data_list = list(queryset)
            for item in data_list:
                if last_timestamp is None:
                    filtered_data.append(item)
                    last_timestamp = item.recorded_at
                else:
                    time_diff = (item.recorded_at - last_timestamp).total_seconds()
                    if time_diff >= interval_seconds:
                        filtered_data.append(item)
                        last_timestamp = item.recorded_at
            queryset = filtered_data
        
        serializer = self.get_serializer(queryset, many=True)
        logger.info(f"Returning {len(serializer.data)} reading records for user_id {user_id}")
        return Response(serializer.data)

@login_required
def logout_confirm(request):
    if request.method == 'POST':
        if 'confirm' in request.POST:
            logout(request)
            return render(request, 'monitoring_app/logout.html')
        return redirect('dashboard')
    return render(request, 'monitoring_app/logout_confirm.html')

@login_required
def live_status(request, user_id):
    user = get_object_or_404(User, id=user_id)
    combined_data = []
    timestamps = set()
    try:
        wq_data = WaterQualityData.objects.filter(user=user).order_by('-timestamp')[:50]
        flow_data = Reading.objects.filter(user=user).order_by('-recorded_at')[:50]
        logger.info(f"Fetched {wq_data.count()} water quality records and {flow_data.count()} flow records for user_id {user_id}")
        
        for item in wq_data:
            timestamps.add(item.timestamp.replace(microsecond=0))
        for item in flow_data:
            timestamps.add(item.recorded_at.replace(microsecond=0))
        
        for ts in sorted(timestamps, reverse=True)[:50]:
            entry = {'timestamp': ts}
            wq_item = next((item for item in wq_data if item.timestamp.replace(microsecond=0) == ts), None)
            if wq_item:
                entry['ph'] = wq_item.ph
                entry['cod'] = wq_item.cod
                entry['bod'] = wq_item.bod
                entry['tss'] = wq_item.tss
            flow_items = [item for item in flow_data if item.recorded_at.replace(microsecond=0) == ts]
            for item in flow_items:
                entry[item.parameter] = item.value
            combined_data.append(entry)
        
        logger.info(f"Combined {len(combined_data)} data entries for user_id {user_id}")
    except Exception as e:
        logger.error(f"Error in live_status for user_id {user_id}: {str(e)}", exc_info=True)
        messages.error(request, "Error loading live status data.")
    
    return render(request, 'monitoring_app/live_status.html', {
        'user': user,
        'combined_data': combined_data
    })

@login_required
def download_page(request, user_id):
    user = get_object_or_404(User, id=user_id)
    return render(request, 'monitoring_app/download.html', {'user': user})

@login_required
def add_data(request):
    users = User.objects.all()
    machines = Machine.objects.all()
    return render(request, 'monitoring_app/add_data.html', {'users': users, 'machines': machines})

@login_required
def data_entry(request):
    if request.user.is_admin:
        users = User.objects.all()
        machines = Machine.objects.all()
        return render(request, 'monitoring_app/data_entry.html', {'users': users, 'machines': machines})
    return redirect('dashboard')

@login_required
def submit_data(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        try:
            user = User.objects.get(id=user_id)
            # Non-flow parameters
            ph = request.POST.get('ph', '') or None
            cod = request.POST.get('cod', '') or None
            bod = request.POST.get('bod', '') or None
            tss = request.POST.get('tss', '') or None
            ph = float(ph) if ph else None
            cod = float(cod) if cod else None
            bod = float(bod) if bod else None
            tss = float(tss) if tss else None
            # Validate non-negative values and upper bounds
            if ph is not None and (ph < 0 or ph > 14):
                return JsonResponse({'success': False, 'error': 'pH must be between 0 and 14'})
            if cod is not None and cod < 0:
                return JsonResponse({'success': False, 'error': 'COD cannot be negative'})
            if bod is not None and bod < 0:
                return JsonResponse({'success': False, 'error': 'BOD cannot be negative'})
            if tss is not None and tss < 0:
                return JsonResponse({'success': False, 'error': 'TSS cannot be negative'})
            # Save non-flow data
            if any([ph, cod, bod, tss]):
                new_wq_data = WaterQualityData(
                    user=user,
                    ph=ph,
                    cod=cod,
                    bod=bod,
                    tss=tss
                )
                new_wq_data.save()
                logger.info(f"Saved water quality data for user_id {user_id}: ph={ph}, cod={cod}, bod={bod}, tss={tss}")
            # Flow parameters
            for flow_number in range(1, 11):
                if getattr(user, f'show_flow{flow_number}'):
                    machine = Machine.objects.get(name=f'machine_flow_{flow_number}')
                    flow_param = f'flow{flow_number}'
                    total_param = f'flow{flow_number}_total'
                    daily_param = f'flow{flow_number}_daily'
                    monthly_param = f'flow{flow_number}_monthly'
                    # Get submitted values
                    flow_value = request.POST.get(flow_param, '') or None
                    total_value = request.POST.get(total_param, '') or None
                    daily_value = request.POST.get(daily_param, '') or None
                    monthly_value = request.POST.get(monthly_param, '') or None
                    flow_value = float(flow_value) if flow_value else None
                    total_value = float(total_value) if total_value else None
                    daily_value = float(daily_value) if daily_value else None
                    monthly_value = float(monthly_value) if monthly_value else None
                    # Validate non-negative values
                    if flow_value is not None and flow_value < 0:
                        return JsonResponse({'success': False, 'error': f'{flow_param} cannot be negative'})
                    if total_value is not None and total_value < 0:
                        return JsonResponse({'success': False, 'error': f'{total_param} cannot be negative'})
                    if daily_value is not None and daily_value < 0:
                        return JsonResponse({'success': False, 'error': f'{daily_param} cannot be negative'})
                    if monthly_value is not None and monthly_value < 0:
                        return JsonResponse({'success': False, 'error': f'{monthly_param} cannot be negative'})
                    # Current date for daily/monthly calculations
                    today = now().date()
                    month_start = today.replace(day=1)
                    # Handle flow value
                    if flow_value is not None:
                        Reading.objects.create(
                            user=user,
                            machine=machine,
                            parameter=flow_param,
                            value=flow_value
                        )
                        logger.info(f"Saved {flow_param} reading with value {flow_value} for user_id {user_id}")
                    # Handle total flow
                    if flow_value is not None or total_value is not None:
                        latest_total = Reading.objects.filter(
                            user=user,
                            machine=machine,
                            parameter=total_param
                        ).order_by('-recorded_at').first()
                        new_total = total_value if total_value is not None else None
                        if new_total is None and flow_value is not None:
                            # Increment total based on flow_value
                            new_total = (latest_total.value if latest_total else 0) + flow_value
                        if new_total is not None:
                            Reading.objects.create(
                                user=user,
                                machine=machine,
                                parameter=total_param,
                                value=new_total
                            )
                            logger.info(f"Saved {total_param} reading with value {new_total} for user_id {user_id}")
                    # Calculate and save daily flow
                    daily_sum = Reading.objects.filter(
                        user=user,
                        machine=machine,
                        parameter=flow_param,
                        recorded_at__date=today
                    ).aggregate(total=Sum('value'))['total'] or 0
                    if flow_value is not None:
                        daily_sum += flow_value
                    if daily_sum > 0 and daily_value is None:  # Only save if there's data and not overridden
                        Reading.objects.create(
                            user=user,
                            machine=machine,
                            parameter=daily_param,
                            value=daily_sum
                        )
                        logger.info(f"Saved {daily_param} reading with value {daily_sum} for user_id {user_id}")
                    elif daily_value is not None:
                        Reading.objects.create(
                            user=user,
                            machine=machine,
                            parameter=daily_param,
                            value=daily_value
                        )
                        logger.info(f"Saved {daily_param} reading with value {daily_value} (manual) for user_id {user_id}")
                    # Calculate and save monthly flow
                    monthly_sum = Reading.objects.filter(
                        user=user,
                        machine=machine,
                        parameter=flow_param,
                        recorded_at__date__gte=month_start
                    ).aggregate(total=Sum('value'))['total'] or 0
                    if flow_value is not None:
                        monthly_sum += flow_value
                    if monthly_sum > 0 and monthly_value is None:  # Only save if there's data and not overridden
                        Reading.objects.create(
                            user=user,
                            machine=machine,
                            parameter=monthly_param,
                            value=monthly_sum
                        )
                        logger.info(f"Saved {monthly_param} reading with value {monthly_sum} for user_id {user_id}")
                    elif monthly_value is not None:
                        Reading.objects.create(
                            user=user,
                            machine=machine,
                            parameter=monthly_param,
                            value=monthly_value
                        )
                        logger.info(f"Saved {monthly_param} reading with value {monthly_value} (manual) for user_id {user_id}")
            return JsonResponse({'success': True})
        except User.DoesNotExist:
            logger.error(f"User not found: user_id {user_id}")
            return JsonResponse({'success': False, 'error': 'User not found'})
        except Machine.DoesNotExist:
            logger.error(f"Machine not found for user_id {user_id}")
            return JsonResponse({'success': False, 'error': 'Machine not found'})
        except ValueError as e:
            logger.error(f"ValueError in submit_data for user_id {user_id}: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
        except Exception as e:
            logger.error(f"Error in submit_data for user_id {user_id}: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'error': str(e)})
    logger.error("Invalid request method in submit_data")
    return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
@login_required
def preview_data(request, user_id):
    if request.method == 'POST':
        try:
            user = User.objects.get(id=user_id)
            body = json.loads(request.body)
            start_date_str = body.get('start_date')
            end_date_str = body.get('end_date')
            time_interval = int(body.get('time_interval', 0))
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            if (end_date - start_date).days > 31:
                return JsonResponse({"error": "Date range cannot exceed 31 days"}, status=400)
            fields = [field for field in ['ph', 'cod', 'bod', 'tss'] +
                      [f'flow{i}' for i in range(1, 11)] +
                      [f'flow{i}_total' for i in range(1, 11)] +
                      [f'flow{i}_daily' for i in range(1, 11)] +
                      [f'flow{i}_monthly' for i in range(1, 11)] if body.get(field)]
            if not fields:
                return JsonResponse({"error": "Please select at least one parameter"}, status=400)
            wq_queryset = WaterQualityData.objects.filter(
                user=user,
                timestamp__date__gte=start_date,
                timestamp__date__lte=end_date
            ).order_by('timestamp').prefetch_related('user')
            flow_queryset = Reading.objects.filter(
                user=user,
                parameter__in=[f'flow{i}' for i in range(1, 11)] +
                              [f'flow{i}_total' for i in range(1, 11)] +
                              [f'flow{i}_daily' for i in range(1, 11)] +
                              [f'flow{i}_monthly' for i in range(1, 11)],
                recorded_at__date__gte=start_date,
                recorded_at__date__lte=end_date
            ).order_by('recorded_at').prefetch_related('user', 'machine')
            if time_interval > 0:
                interval_seconds = time_interval * 60
                filtered_wq = []
                filtered_flow = []
                last_wq_timestamp = last_flow_timestamp = None
                for item in wq_queryset:
                    if last_wq_timestamp is None or (item.timestamp - last_wq_timestamp).total_seconds() >= interval_seconds:
                        filtered_wq.append(item)
                        last_wq_timestamp = item.timestamp
                for item in flow_queryset:
                    if last_flow_timestamp is None or (item.recorded_at - last_flow_timestamp).total_seconds() >= interval_seconds:
                        filtered_flow.append(item)
                        last_flow_timestamp = item.recorded_at
                wq_queryset = filtered_wq
                flow_queryset = filtered_flow
            combined_data = {}
            wq_data = list(wq_queryset)
            flow_data = list(flow_queryset)
            for item in wq_data:
                rounded_ts = item.timestamp.replace(microsecond=0)
                if rounded_ts not in combined_data:
                    combined_data[rounded_ts] = {'timestamp': rounded_ts.isoformat()}
                if 'ph' in fields:
                    combined_data[rounded_ts]['ph'] = item.ph
                if 'cod' in fields:
                    combined_data[rounded_ts]['cod'] = item.cod
                if 'bod' in fields:
                    combined_data[rounded_ts]['bod'] = item.bod
                if 'tss' in fields:
                    combined_data[rounded_ts]['tss'] = item.tss
            for item in flow_data:
                rounded_ts = item.recorded_at.replace(microsecond=0)
                if rounded_ts not in combined_data:
                    combined_data[rounded_ts] = {'timestamp': rounded_ts.isoformat()}
                if item.parameter in fields:
                    combined_data[rounded_ts][item.parameter] = item.value
            combined_data_list = sorted(combined_data.values(), key=lambda x: x['timestamp'])
            if not combined_data_list:
                return JsonResponse({"error": "No data available for the selected parameters and date range"}, status=400)
            return JsonResponse({
                'fields': fields,
                'data': combined_data_list
            })
        except User.DoesNotExist:
            return JsonResponse({"error": "User not found"}, status=404)
        except ValueError as e:
            return JsonResponse({"error": f"Invalid input: {str(e)}"}, status=400)
        except Exception as e:
            logger.error(f"Error in preview_data: {e}")
            return JsonResponse({"error": f"Error generating preview: {str(e)}"}, status=500)
    return JsonResponse({"error": "Invalid request method"}, status=400)

@login_required
def download_data(request, user_id):
    if request.method == 'POST':
        try:
            user = User.objects.get(id=user_id)
            body = json.loads(request.body)
            start_date_str = body.get('start_date')
            end_date_str = body.get('end_date')
            file_format = body.get('format')
            time_interval = int(body.get('time_interval', 0))
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            if (end_date - start_date).days > 31:
                return JsonResponse({"error": "Date range cannot exceed 31 days"}, status=400)
            fields = [field for field in ['ph', 'cod', 'bod', 'tss'] +
                      [f'flow{i}' for i in range(1, 11)] +
                      [f'flow{i}_total' for i in range(1, 11)] +
                      [f'flow{i}_daily' for i in range(1, 11)] +
                      [f'flow{i}_monthly' for i in range(1, 11)] if body.get(field)]
            if not fields:
                return JsonResponse({"error": "Please select at least one parameter"}, status=400)
            wq_queryset = WaterQualityData.objects.filter(
                user=user,
                timestamp__date__gte=start_date,
                timestamp__date__lte=end_date
            ).order_by('timestamp').prefetch_related('user')
            flow_queryset = Reading.objects.filter(
                user=user,
                parameter__in=[f'flow{i}' for i in range(1, 11)] +
                              [f'flow{i}_total' for i in range(1, 11)] +
                              [f'flow{i}_daily' for i in range(1, 11)] +
                              [f'flow{i}_monthly' for i in range(1, 11)],
                recorded_at__date__gte=start_date,
                recorded_at__date__lte=end_date
            ).order_by('recorded_at').prefetch_related('user', 'machine')
            if time_interval > 0:
                interval_seconds = time_interval * 60
                filtered_wq = []
                filtered_flow = []
                last_wq_timestamp = last_flow_timestamp = None
                for item in wq_queryset:
                    if last_wq_timestamp is None or (item.timestamp - last_wq_timestamp).total_seconds() >= interval_seconds:
                        filtered_wq.append(item)
                        last_wq_timestamp = item.timestamp
                for item in flow_queryset:
                    if last_flow_timestamp is None or (item.recorded_at - last_flow_timestamp).total_seconds() >= interval_seconds:
                        filtered_flow.append(item)
                        last_flow_timestamp = item.recorded_at
                wq_queryset = filtered_wq
                flow_queryset = filtered_flow
            combined_data = {}
            wq_data = list(wq_queryset)
            flow_data = list(flow_queryset)
            for item in wq_data:
                rounded_ts = item.timestamp.replace(microsecond=0)
                if rounded_ts not in combined_data:
                    combined_data[rounded_ts] = {'timestamp': rounded_ts}
                if 'ph' in fields:
                    combined_data[rounded_ts]['ph'] = item.ph
                if 'cod' in fields:
                    combined_data[rounded_ts]['cod'] = item.cod
                if 'bod' in fields:
                    combined_data[rounded_ts]['bod'] = item.bod
                if 'tss' in fields:
                    combined_data[rounded_ts]['tss'] = item.tss
            for item in flow_data:
                rounded_ts = item.recorded_at.replace(microsecond=0)
                if rounded_ts not in combined_data:
                    combined_data[rounded_ts] = {'timestamp': rounded_ts}
                if item.parameter in fields:
                    combined_data[rounded_ts][item.parameter] = item.value
            combined_data_list = sorted(combined_data.values(), key=lambda x: x['timestamp'])
            if not combined_data_list:
                return JsonResponse({"error": "No data available for the selected parameters and date range"}, status=400)
            if file_format == 'excel':
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Water Quality Data"
                headers = ['Date/Time'] + [f"{field.upper()} ({'L/min' if field.startswith('flow') and not any(s in field for s in ['total', 'daily', 'monthly']) else 'L' if field.startswith('flow') else 'mg/L' if field in ['cod', 'bod', 'tss'] else ''})" for field in fields]
                ws.append(headers)
                for row in combined_data_list:
                    row_data = [row['timestamp'].strftime('%Y-%m-%d %H:%M:%S')]
                    for field in fields:
                        value = row.get(field)
                        row_data.append(round(value, 2) if value is not None else '')
                    ws.append(row_data)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for column_cells in ws.columns:
                    max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                    ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="water_quality_data_{user.username}_{start_date}_{end_date}.xlsx"'
                response.write(buffer.getvalue())
                return response
            elif file_format == 'pdf':
                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
                elements = []
                styles = getSampleStyleSheet()
                elements.append(Paragraph(f"Water Quality Data for {user.username}", styles['Title']))
                elements.append(Spacer(1, 12))
                headers = ['Date/Time'] + [f"{field.upper()} ({'L/min' if field.startswith('flow') and not any(s in field for s in ['total', 'daily', 'monthly']) else 'L' if field.startswith('flow') else 'mg/L' if field in ['cod', 'bod', 'tss'] else ''})" for field in fields]
                data = [headers]
                for row in combined_data_list:
                    row_data = [row['timestamp'].strftime('%Y-%m-%d %H:%M:%S')]
                    for field in fields:
                        value = row.get(field)
                        row_data.append(f"{value:.2f}" if value is not None else '')
                    data.append(row_data)
                num_columns = len(headers)
                page_width = landscape(letter)[0] - 72
                col_width = page_width / num_columns
                col_widths = [col_width] * num_columns
                table = Table(data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('WORDWRAP', (0, 0), (-1, -1), True),
                ]))
                elements.append(table)
                doc.build(elements)
                pdf = buffer.getvalue()
                buffer.close()
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="water_quality_data_{user.username}_{start_date}_{end_date}.pdf"'
                response.write(pdf)
                return response
        except User.DoesNotExist:
            return JsonResponse({"error": "User not found"}, status=404)
        except ValueError as e:
            return JsonResponse({"error": f"Invalid input: {str(e)}"}, status=400)
        except Exception as e:
            logger.error(f"Error in download_data: {e}")
            return JsonResponse({"error": f"Error generating file: {str(e)}"}, status=500)
    return JsonResponse({"error": "Invalid request method"}, status=400)